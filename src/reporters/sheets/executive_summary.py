"""Executive summary sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict, List
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class ExecutiveSummarySheet(BaseSheet):
    """Sheet generator for executive summary."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, metrics: Dict[str, Any], web_acls: List[Dict[str, Any]]) -> None:
        """
        Create the Executive Summary sheet with Web ACL information.
        """
        logger.info("Creating Executive Summary sheet...")

        ws = self.workbook.create_sheet("Executive Summary", 0)

        # Title with professional styling
        ws['A1'] = 'AWS WAF Security Analysis - Executive Summary'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:D2')

        row = 4

        # Web ACLs Summary Section
        ws[f'A{row}'] = 'Web ACLs Overview'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # Web ACL details
        for acl in web_acls:
            acl_name = acl.get('name', acl.get('Name', ''))
            acl_scope = acl.get('scope', acl.get('Scope', ''))
            acl_capacity = acl.get('capacity', acl.get('Capacity', 0))
            default_action = acl.get('default_action', acl.get('DefaultAction', {}))

            if isinstance(default_action, str):
                action_str = default_action
            else:
                action_str = 'ALLOW' if 'Allow' in default_action else 'BLOCK'

            ws[f'A{row}'] = f"• {acl_name}"
            ws[f'A{row}'].font = Font(bold=True, size=11, name='Calibri')
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            # Format ACL details with borders and consistent styling
            acl_details = [
                ("  Scope:", acl_scope),
                ("  Default Action:", action_str),
                ("  Capacity Used:", f"{acl_capacity} WCU")
            ]

            for label, value in acl_details:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.data_font
                ws[f'A{row}'].border = self.thin_border
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'B{row}'].border = self.thin_border
                row += 1

            row += 1  # Extra space between ACLs

        # Summary metrics
        summary = metrics.get('summary', {})
        coverage = metrics.get('web_acl_coverage', {})

        row += 1
        ws[f'A{row}'] = 'Key Security Metrics'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        key_metrics = [
            ('Total Requests Analyzed', f"{summary.get('total_requests', 0):,}"),
            ('Blocked Requests', f"{summary.get('blocked_requests', 0):,}"),
            ('Block Rate', f"{summary.get('block_rate_percent', 0):.2f}%"),
            ('Unique Client IPs', f"{summary.get('unique_client_ips', 0):,}"),
            ('Unique Countries', f"{summary.get('unique_countries', 0):,}"),
            ('Web ACLs Configured', coverage.get('total_web_acls', 0)),
            ('Protected Resources', coverage.get('total_protected_resources', 0)),
            ('Logging Coverage', f"{coverage.get('logging_coverage_percent', 0):.1f}%")
        ]

        for idx, (metric_name, metric_value) in enumerate(key_metrics):
            highlight = idx % 2 == 0

            ws[f'A{row}'] = metric_name
            ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
            ws[f'A{row}'].border = self.thin_border
            if highlight:
                ws[f'A{row}'].fill = self.highlight_fill

            ws[f'B{row}'] = metric_value
            ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E78', name='Calibri')
            ws[f'B{row}'].border = self.thin_border
            if highlight:
                ws[f'B{row}'].fill = self.highlight_fill
            ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')

            row += 1

        # Time range
        row += 1
        ws[f'A{row}'] = 'Analysis Period'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        time_range = summary.get('time_range')
        if time_range:
            time_data = [
                ('Start Date', str(time_range.get('start', ''))[:19]),
                ('End Date', str(time_range.get('end', ''))[:19])
            ]

            for label, value in time_data:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'A{row}'].border = self.thin_border
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = self.data_font
                ws[f'B{row}'].border = self.thin_border
                row += 1

        # Security Posture Score
        security_score = coverage.get('security_posture_score', 0)
        if security_score:
            row += 1
            ws[f'A{row}'] = 'Security Posture Score'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            ws[f'A{row}'] = 'Overall Score'
            ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
            ws[f'A{row}'].border = self.thin_border

            ws[f'B{row}'] = f"{security_score}/100"
            ws[f'B{row}'].border = self.thin_border
            ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')

            # Color code the score with appropriate fill
            if security_score >= 80:
                ws[f'B{row}'].font = Font(bold=True, size=14, color='008000', name='Calibri')  # Green
                ws[f'B{row}'].fill = self.success_fill
            elif security_score >= 60:
                ws[f'B{row}'].font = Font(bold=True, size=14, color='FF8C00', name='Calibri')  # Orange
                ws[f'B{row}'].fill = self.warning_fill
            else:
                ws[f'B{row}'].font = Font(bold=True, size=14, color='FF0000', name='Calibri')  # Red
                ws[f'B{row}'].fill = self.danger_fill
            row += 1

        # Action distribution chart
        row += 2
        action_dist = metrics.get('action_distribution', {})
        if action_dist:
            try:
                chart_buffer = self.viz.create_action_distribution_chart(action_dist)
                img = XLImage(chart_buffer)
                img.width = 600
                img.height = 400
                ws.add_image(img, f'A{row}')
            except Exception as e:
                logger.warning(f"Could not create action distribution chart: {e}")

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
