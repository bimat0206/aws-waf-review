"""Inventory sheet generator."""

import json
import logging
from datetime import datetime
from typing import Any, Dict, List
from openpyxl.styles import Alignment, Font
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class InventorySheet(BaseSheet):
    """Sheet generator for inventory."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, web_acls: List[Dict[str, Any]],
                               resources: List[Dict[str, Any]],
                               logging_configs: List[Dict[str, Any]],
                               rules_by_web_acl: Dict[str, List[Dict[str, Any]]]) -> None:
        """
        Create the Inventory sheet with Web ACLs, resources, and rules.
        """
        logger.info("Creating Inventory sheet...")

        ws = self.workbook.create_sheet("Inventory")

        # Title with professional styling
        ws['A1'] = 'AWS WAF Inventory'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 30

        # Add a subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:H2')

        # Web ACLs section
        row = 4
        ws[f'A{row}'] = 'Web ACLs Configuration'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        # Headers with professional formatting
        headers = ['Name', 'ID', 'Scope', 'Default Action', 'Capacity', 'Rules Count', 'Resources Count', 'Logging']
        self._format_header_row(ws, row, headers)
        row += 1

        # Data
        logging_config_ids = {lc.get('web_acl_id') for lc in logging_configs if lc.get('web_acl_id')}
        resources_count = {}
        for resource in resources:
            web_acl_id = resource.get('web_acl_id')
            if web_acl_id:
                resources_count[web_acl_id] = resources_count.get(web_acl_id, 0) + 1

        for idx, acl in enumerate(web_acls):
            acl_id = acl.get('web_acl_id', acl.get('Id', ''))
            acl_name = acl.get('name', acl.get('Name', ''))

            # Apply alternating row colors
            highlight = idx % 2 == 0

            default_action = acl.get('default_action', acl.get('DefaultAction', {}))
            if isinstance(default_action, str):
                action_str = default_action
            else:
                action_str = 'ALLOW' if 'Allow' in default_action else 'BLOCK'

            rules_count = len(rules_by_web_acl.get(acl_id, []))
            logging_enabled = 'Yes' if acl_id in logging_config_ids else 'No'

            # Apply professional styling to all cells in the row
            row_data = [
                acl_name,
                acl_id,
                acl.get('scope', acl.get('Scope', '')),
                action_str,
                acl.get('capacity', acl.get('Capacity', 0)),
                rules_count,
                resources_count.get(acl_id, 0),
                logging_enabled
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx)
                self._format_data_cell(cell, value, highlight)

            # Special color coding for logging status
            logging_cell = ws.cell(row=row, column=8)
            if logging_enabled == 'No':
                logging_cell.fill = self.danger_fill
            else:
                logging_cell.fill = self.success_fill

            row += 1

        # Rule Implementation Summary section
        row += 2
        ws[f'A{row}'] = 'Rule Implementation Summary'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

        # Calculate rule statistics
        total_rules = sum(len(rules) for rules in rules_by_web_acl.values())
        rule_types = {}
        rule_actions = {}

        for web_acl_id, rules in rules_by_web_acl.items():
            for rule in rules:
                # Count by type
                rule_type = rule.get('rule_type', 'UNKNOWN')
                rule_types[rule_type] = rule_types.get(rule_type, 0) + 1

                # Count by action
                action = rule.get('action', '')
                if isinstance(action, str) and action:
                    try:
                        action_dict = json.loads(action)
                        if 'Allow' in action_dict:
                            action_key = 'ALLOW'
                        elif 'Block' in action_dict:
                            action_key = 'BLOCK'
                        elif 'Count' in action_dict:
                            action_key = 'COUNT'
                        elif 'Captcha' in action_dict:
                            action_key = 'CAPTCHA'
                        elif 'Challenge' in action_dict:
                            action_key = 'CHALLENGE'
                        else:
                            action_key = 'OTHER'
                        rule_actions[action_key] = rule_actions.get(action_key, 0) + 1
                    except:
                        rule_actions['UNKNOWN'] = rule_actions.get('UNKNOWN', 0) + 1

        # Summary statistics
        ws[f'A{row}'] = 'Total Rules Configured'
        ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
        ws[f'A{row}'].border = self.thin_border
        ws[f'B{row}'] = total_rules
        ws[f'B{row}'].font = Font(bold=True, size=11, color='1F4E78', name='Calibri')
        ws[f'B{row}'].border = self.thin_border
        ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
        row += 1

        # Rules by type
        if rule_types:
            row += 1
            ws[f'A{row}'] = 'Rules by Type:'
            ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

            headers = ['Rule Type', 'Count', 'Percentage']
            self._format_header_row(ws, row, headers, start_col=1)
            row += 1

            sorted_types = sorted(rule_types.items(), key=lambda x: x[1], reverse=True)
            for idx, (rule_type, count) in enumerate(sorted_types):
                highlight = idx % 2 == 0
                percentage = (count / total_rules * 100) if total_rules > 0 else 0

                row_data = [rule_type, count, f"{percentage:.1f}%"]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)
                row += 1

        # Rules by action
        if rule_actions:
            row += 1
            ws[f'A{row}'] = 'Rules by Action:'
            ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
            ws.merge_cells(f'A{row}:B{row}')
            row += 1

            headers = ['Action', 'Count', 'Percentage']
            self._format_header_row(ws, row, headers, start_col=1)
            row += 1

            sorted_actions = sorted(rule_actions.items(), key=lambda x: x[1], reverse=True)
            for idx, (action, count) in enumerate(sorted_actions):
                highlight = idx % 2 == 0
                percentage = (count / total_rules * 100) if total_rules > 0 else 0

                row_data = [action, count, f"{percentage:.1f}%"]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                # Color code based on action
                action_cell = ws.cell(row=row, column=1)
                if action == 'BLOCK':
                    action_cell.font = Font(bold=True, size=10, color='C00000', name='Calibri')
                elif action == 'ALLOW':
                    action_cell.font = Font(bold=True, size=10, color='008000', name='Calibri')
                elif action in ['CAPTCHA', 'CHALLENGE']:
                    action_cell.font = Font(bold=True, size=10, color='FF8C00', name='Calibri')

                row += 1

        # Rules section (detailed view)
        row += 2
        ws[f'A{row}'] = 'Rules and Rule Groups'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:E{row}')
        row += 1

        # Headers
        headers = ['Web ACL', 'Rule Name', 'Priority', 'Type', 'Action']
        self._format_header_row(ws, row, headers)
        row += 1

        # Get Web ACL names mapping
        web_acl_names = {acl.get('web_acl_id', acl.get('Id', '')): acl.get('name', acl.get('Name', '')) for acl in web_acls}

        # Data - rules sorted by Web ACL and priority
        all_rules = []
        for web_acl_id, rules in rules_by_web_acl.items():
            for rule in rules:
                all_rules.append({
                    'web_acl_name': web_acl_names.get(web_acl_id, web_acl_id),
                    'rule': rule
                })

        # Sort by Web ACL name and priority
        all_rules.sort(key=lambda x: (x['web_acl_name'], x['rule'].get('priority', 0)))

        for idx, item in enumerate(all_rules):
            rule = item['rule']

            # Parse action from JSON string if needed
            action = rule.get('action', '')
            if isinstance(action, str):
                try:
                    action_dict = json.loads(action) if action else {}
                    if 'Allow' in action_dict:
                        action = 'ALLOW'
                    elif 'Block' in action_dict:
                        action = 'BLOCK'
                    elif 'Count' in action_dict:
                        action = 'COUNT'
                    elif 'Captcha' in action_dict:
                        action = 'CAPTCHA'
                    elif 'Challenge' in action_dict:
                        action = 'CHALLENGE'
                except:
                    pass

            # Apply alternating row colors
            highlight = idx % 2 == 0

            row_data = [
                item['web_acl_name'],
                rule.get('name', ''),
                rule.get('priority', ''),
                rule.get('rule_type', ''),
                action
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx)
                self._format_data_cell(cell, value, highlight)

            row += 1

        # Resources section
        row += 2
        ws[f'A{row}'] = 'Protected Resources'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:C{row}')
        row += 1

        # Headers
        headers = ['Web ACL Name', 'Resource Type', 'Resource ARN']
        self._format_header_row(ws, row, headers)
        row += 1

        # Data
        if resources:
            for idx, resource in enumerate(resources):
                highlight = idx % 2 == 0

                row_data = [
                    resource.get('web_acl_name', ''),
                    resource.get('resource_type', ''),
                    resource.get('resource_arn', '')
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                row += 1
        else:
            ws[f'A{row}'] = 'No resources associated with Web ACLs'
            ws[f'A{row}'].font = Font(italic=True, color='808080', name='Calibri')
            ws[f'A{row}'].alignment = Alignment(vertical='center')
            ws.merge_cells(f'A{row}:C{row}')
            row += 1

        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
