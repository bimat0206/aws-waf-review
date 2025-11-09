"""Client analysis sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class ClientAnalysisSheet(BaseSheet):
    """Sheet generator for client analysis."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, metrics: Dict[str, Any]) -> None:
        """
        Create the Client Analysis sheet.

        This sheet analyzes client behavior and identifies potential threats, including:
        - Top blocked IP addresses and their activity patterns
        - Bot traffic analysis using JA3/JA4 fingerprints
        - User agent analysis to identify automated tools
        - Hourly traffic patterns to detect anomalies
        - Temporal analysis of repeat offenders

        Metrics explained:
        - Block Count: Total blocks for each IP address
        - Unique Rules Hit: Number of different rules triggered by the IP
        - First/Last Seen: Time range of activity (useful for identifying ongoing attacks)
        - JA3/JA4 Fingerprints: TLS fingerprinting for bot detection
        - User Agents: Client identification strings (can reveal malicious tools)
        """
        logger.info("Creating Client Analysis sheet...")

        ws = self.workbook.create_sheet("Client Analysis")

        # Title with professional styling
        ws['A1'] = 'Client and Bot Analysis'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:F2')

        # Description
        ws['A3'] = 'Analyzes client behavior, identifies malicious IP addresses, and detects bot traffic to enhance threat detection and response.'
        ws['A3'].font = Font(size=10, italic=True, color='606060', name='Calibri')
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:F3')
        ws.row_dimensions[3].height = 30

        row = 5

        # Top blocked IPs
        top_ips = metrics.get('top_blocked_ips', [])
        if top_ips:
            ws[f'A{row}'] = 'Top Blocked IP Addresses'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Headers
            headers = ['IP Address', 'Country', 'Block Count', 'Unique Rules Hit', 'First Seen', 'Last Seen']
            self._format_header_row(ws, row, headers)
            row += 1

            # Data
            for idx, ip_data in enumerate(top_ips[:30]):  # Top 30
                highlight = idx % 2 == 0

                row_data = [
                    ip_data.get('ip', ''),
                    ip_data.get('country', ''),
                    ip_data.get('block_count', 0),
                    ip_data.get('unique_rules_hit', 0),
                    str(ip_data.get('first_seen', ''))[:19],
                    str(ip_data.get('last_seen', ''))[:19]
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                row += 1

        # Bot analysis
        row += 2
        bot_analysis = metrics.get('bot_analysis', {})
        if bot_analysis:
            ws[f'A{row}'] = 'Bot Traffic Analysis'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            # Bot metrics with professional styling
            bot_metrics = [
                ('Requests with JA3 Fingerprint', bot_analysis.get('requests_with_ja3', 0)),
                ('Requests with JA4 Fingerprint', bot_analysis.get('requests_with_ja4', 0))
            ]

            for idx, (label, value) in enumerate(bot_metrics):
                highlight = idx % 2 == 0
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'A{row}'].border = self.thin_border
                if highlight:
                    ws[f'A{row}'].fill = self.highlight_fill

                ws[f'B{row}'] = value
                ws[f'B{row}'].font = self.data_font
                ws[f'B{row}'].border = self.thin_border
                if highlight:
                    ws[f'B{row}'].fill = self.highlight_fill
                row += 1

            row += 1

            # Top user agents
            top_agents = bot_analysis.get('top_user_agents', [])
            if top_agents:
                ws[f'A{row}'] = 'Top User Agents'
                ws[f'A{row}'].font = Font(bold=True, size=11, name='Calibri')
                ws.merge_cells(f'A{row}:B{row}')
                row += 1

                # Headers
                headers = ['User Agent', 'Request Count']
                self._format_header_row(ws, row, headers, start_col=1)
                row += 1

                for idx, agent_data in enumerate(top_agents[:15]):
                    highlight = idx % 2 == 0
                    row_data = [
                        agent_data.get('user_agent', '')[:100],
                        agent_data.get('count', 0)
                    ]

                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        self._format_data_cell(cell, value, highlight)

                    row += 1

        # Hourly patterns chart - positioned on the right side
        hourly_data = metrics.get('hourly_patterns', [])
        if hourly_data:
            try:
                chart_buffer = self.viz.create_hourly_pattern_chart(hourly_data)
                img = XLImage(chart_buffer)
                img.width = 700
                img.height = 400
                # Place chart starting at column H (right side of the data)
                ws.add_image(img, 'H5')
            except Exception as e:
                logger.warning(f"Could not create hourly pattern chart: {e}")

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['G'].width = 2  # Gap
        ws.column_dimensions['H'].width = 2
        ws.column_dimensions['I'].width = 2
        ws.column_dimensions['J'].width = 2
        ws.column_dimensions['K'].width = 2
