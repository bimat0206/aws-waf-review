"""Geographic blocked traffic sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict, List
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class GeographicBlockedTrafficSheet(BaseSheet):
    """Sheet generator for geographic."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, metrics: Dict[str, Any]) -> None:
        """
        Create the Geographic Distribution of Blocked Traffic sheet.

        This sheet focuses exclusively on blocked traffic by geographic location, including:
        - Countries with the highest volume of blocked requests
        - Block rates by country
        - Threat level assessments based on block patterns
        - Risk assessments for each region
        - Visual geographic threat heatmap

        Metrics explained:
        - Blocked Requests: Total requests blocked from this country
        - Total Requests: All requests from this country (blocked + allowed)
        - Block Rate %: Percentage of requests blocked
        - Threat Level: Assessment based on volume and block rate
          - CRITICAL: >75% block rate with >100 blocked requests
          - HIGH: >50% block rate with >50 blocked requests
          - MEDIUM: >25% block rate or >100 blocked requests
          - LOW: Below medium thresholds
        """
        logger.info("Creating Geographic Distribution of Blocked Traffic sheet...")

        ws = self.workbook.create_sheet("Geographic Blocked Traffic")

        # Title with professional styling
        ws['A1'] = 'Geographic Distribution of Blocked Traffic'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:F2')

        # Description
        ws['A3'] = 'Identifies geographic origins of blocked requests to help identify regions that may be sources of malicious traffic or targeted attacks.'
        ws['A3'].font = Font(size=10, italic=True, color='606060', name='Calibri')
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:F3')
        ws.row_dimensions[3].height = 30

        row = 5

        # Get geographic data focused on blocked traffic
        geo_data = metrics.get('geographic_distribution', [])

        if geo_data:
            # Filter and sort by blocked requests
            blocked_geo_data = [
                {**country, 'block_rate': (country.get('blocked_requests', 0) / country.get('total_requests', 1) * 100)
                    if country.get('total_requests', 0) > 0 else 0}
                for country in geo_data
                if country.get('blocked_requests', 0) > 0
            ]
            blocked_geo_data.sort(key=lambda x: x.get('blocked_requests', 0), reverse=True)

            # Summary statistics
            total_blocked = sum(c.get('blocked_requests', 0) for c in blocked_geo_data)
            total_countries_with_blocks = len(blocked_geo_data)

            ws[f'A{row}'] = 'Blocked Traffic Summary'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            summary_data = [
                ('Total Blocked Requests', f"{total_blocked:,}"),
                ('Countries with Blocked Traffic', total_countries_with_blocks),
                ('Top Blocking Country', blocked_geo_data[0].get('country', 'N/A') if blocked_geo_data else 'N/A'),
                ('Top Country Blocked Requests', f"{blocked_geo_data[0].get('blocked_requests', 0):,}" if blocked_geo_data else '0')
            ]

            for idx, (label, value) in enumerate(summary_data):
                highlight = idx % 2 == 0
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'A{row}'].border = self.thin_border
                if highlight:
                    ws[f'A{row}'].fill = self.highlight_fill

                ws[f'B{row}'] = value
                ws[f'B{row}'].font = Font(bold=True, size=10, color='1F4E78', name='Calibri')
                ws[f'B{row}'].border = self.thin_border
                if highlight:
                    ws[f'B{row}'].fill = self.highlight_fill
                ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
                row += 1

            # Detailed table
            row += 2
            ws[f'A{row}'] = 'Blocked Traffic by Country (Top 30)'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1

            headers = ['Country', 'Blocked Requests', 'Total Requests', 'Block Rate %', 'Threat Level', 'Risk Assessment']
            self._format_header_row(ws, row, headers)
            row += 1

            for idx, country_data in enumerate(blocked_geo_data[:30]):
                highlight = idx % 2 == 0
                blocked = country_data.get('blocked_requests', 0)
                total = country_data.get('total_requests', 0)
                block_rate = country_data.get('block_rate', 0)

                # Determine threat level based on block rate and volume
                if block_rate > 75 and blocked > 100:
                    threat_level = 'CRITICAL'
                    risk_assessment = 'High volume of blocked traffic - investigate immediately'
                    threat_fill = self.danger_fill
                    threat_color = 'C00000'
                elif block_rate > 50 and blocked > 50:
                    threat_level = 'HIGH'
                    risk_assessment = 'Significant blocking activity - monitor closely'
                    threat_fill = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')
                    threat_color = 'FF6600'
                elif block_rate > 25 or blocked > 100:
                    threat_level = 'MEDIUM'
                    risk_assessment = 'Moderate threat activity detected'
                    threat_fill = self.warning_fill
                    threat_color = 'FF8C00'
                else:
                    threat_level = 'LOW'
                    risk_assessment = 'Low threat activity'
                    threat_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                    threat_color = '0066CC'

                row_data = [
                    country_data.get('country', ''),
                    blocked,
                    total,
                    f"{block_rate:.1f}%",
                    threat_level,
                    risk_assessment
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                # Color code threat level column
                threat_cell = ws.cell(row=row, column=5)
                threat_cell.fill = threat_fill
                threat_cell.font = Font(bold=True, size=10, color=threat_color, name='Calibri')

                row += 1

        else:
            ws[f'A{row}'] = 'No geographic data available for blocked traffic analysis'
            ws[f'A{row}'].font = Font(italic=True, color='808080', name='Calibri')
            ws.merge_cells(f'A{row}:F{row}')

        # Add visualization on the right side
        if geo_data:
            try:
                chart_buffer = self.viz.create_geographic_threat_chart(geo_data)
                img = XLImage(chart_buffer)
                img.width = 700
                img.height = 450
                # Place chart starting at column H (right side of the data table)
                ws.add_image(img, 'H5')
            except Exception as e:
                logger.warning(f"Could not create geographic chart: {e}")

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 2  # Gap
        ws.column_dimensions['H'].width = 2
        ws.column_dimensions['I'].width = 2
        ws.column_dimensions['J'].width = 2
        ws.column_dimensions['K'].width = 2
