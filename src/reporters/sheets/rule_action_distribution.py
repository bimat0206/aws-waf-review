"""Rule action distribution sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class RuleActionDistributionSheet(BaseSheet):
    """Sheet generator for rule action."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, metrics: Dict[str, Any], llm_findings: list = None) -> None:
        """
        Create the Rule Action Distribution sheet.
        Analyzes rule actions (BLOCK, ALLOW, CHALLENGE, CAPTCHA) to evaluate rule effectiveness.

        Args:
            metrics: Dictionary containing calculated metrics
            llm_findings: Optional list of LLM-generated findings
        """
        logger.info("Creating Rule Action Distribution sheet...")

        ws = self.workbook.create_sheet("Rule Action Distribution")

        # Title with professional styling
        ws['A1'] = 'Rule Action Distribution Analysis'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:G2')

        # Description
        ws['A3'] = 'Analyzes AWS WAF rule actions to evaluate effectiveness, identify imbalances, and determine if adjustments are needed to enhance security without impacting legitimate traffic.'
        ws['A3'].font = Font(size=10, italic=True, color='606060', name='Calibri')
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:G3')
        ws.row_dimensions[3].height = 30

        row = 5

        # Get action distribution from metrics
        action_dist = metrics.get('action_distribution', {})
        rule_effectiveness = metrics.get('rule_effectiveness', [])

        if action_dist:
            # Overall action distribution summary
            ws[f'A{row}'] = 'Overall Action Distribution'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:G{row}')
            row += 1

            def _extract_count(value: Any) -> float:
                if isinstance(value, dict):
                    return value.get('count', 0)
                return value or 0

            total_actions = sum(_extract_count(v) for v in action_dist.values())

            # Summary table
            headers = ['Action', 'Count', 'Percentage', 'Security Impact', 'Recommendation']
            self._format_header_row(ws, row, headers)
            row += 1

            action_order = ['BLOCK', 'ALLOW', 'COUNT', 'CAPTCHA', 'CHALLENGE']
            for idx, action in enumerate(action_order):
                if action in action_dist:
                    highlight = idx % 2 == 0
                    count = _extract_count(action_dist[action])
                    percentage = (count / total_actions * 100) if total_actions > 0 else 0

                    # Determine security impact and recommendation
                    if action == 'BLOCK':
                        impact = 'High Security'
                        recommendation = f'{percentage:.1f}% of traffic blocked - ensure legitimate traffic is not impacted'
                        impact_color = '008000'
                    elif action == 'ALLOW':
                        impact = 'Low Security'
                        recommendation = f'{percentage:.1f}% allowed - verify rules are properly configured'
                        impact_color = 'FF8C00'
                    elif action == 'COUNT':
                        impact = 'Monitoring Only'
                        recommendation = 'Consider converting to BLOCK if threats are confirmed'
                        impact_color = '0066CC'
                    elif action in ['CAPTCHA', 'CHALLENGE']:
                        impact = 'Medium Security'
                        recommendation = 'Good balance of security and user experience'
                        impact_color = '6BCF7F'
                    else:
                        impact = 'Unknown'
                        recommendation = 'Review rule configuration'
                        impact_color = '808080'

                    row_data = [action, count, f"{percentage:.1f}%", impact, recommendation]

                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row, column=col_idx)
                        self._format_data_cell(cell, value, highlight)

                    # Color code action column
                    action_cell = ws.cell(row=row, column=1)
                    if action == 'BLOCK':
                        action_cell.font = Font(bold=True, size=10, color='C00000', name='Calibri')
                        action_cell.fill = self.danger_fill
                    elif action == 'ALLOW':
                        action_cell.font = Font(bold=True, size=10, color='008000', name='Calibri')
                        action_cell.fill = self.success_fill
                    elif action in ['CAPTCHA', 'CHALLENGE']:
                        action_cell.font = Font(bold=True, size=10, color='FF8C00', name='Calibri')
                        action_cell.fill = self.warning_fill

                    # Color code impact column
                    impact_cell = ws.cell(row=row, column=4)
                    impact_cell.font = Font(bold=True, size=10, color=impact_color, name='Calibri')

                    row += 1

        # Rule-level action analysis
        if rule_effectiveness:
            row += 2
            ws[f'A{row}'] = 'Rule-Level Action Analysis'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:G{row}')
            row += 1

            headers = ['Rule ID', 'Total Hits', 'Blocks', 'Allows', 'Block Rate %', 'Effectiveness', 'Status']
            self._format_header_row(ws, row, headers)
            row += 1

            for idx, rule in enumerate(rule_effectiveness[:50]):  # Top 50 rules
                highlight = idx % 2 == 0

                hit_count = rule.get('hit_count', 0)
                blocks = rule.get('blocks', 0)
                allows = rule.get('allows', 0)
                block_rate = rule.get('block_rate_percent', 0)

                # Determine effectiveness
                if hit_count == 0:
                    effectiveness = 'UNUSED'
                    status = 'Consider removing or reviewing'
                    status_fill = self.danger_fill
                elif block_rate > 80:
                    effectiveness = 'HIGHLY EFFECTIVE'
                    status = 'Performing well'
                    status_fill = self.success_fill
                elif block_rate > 50:
                    effectiveness = 'EFFECTIVE'
                    status = 'Good performance'
                    status_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                elif block_rate > 20:
                    effectiveness = 'MODERATE'
                    status = 'Review for optimization'
                    status_fill = self.warning_fill
                else:
                    effectiveness = 'LOW'
                    status = 'May need adjustment'
                    status_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')

                row_data = [
                    rule.get('rule_id', '')[:40],
                    hit_count,
                    blocks,
                    allows,
                    f"{block_rate:.1f}%",
                    effectiveness,
                    status
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                # Color code effectiveness column
                eff_cell = ws.cell(row=row, column=6)
                eff_cell.fill = status_fill
                eff_cell.font = Font(bold=True, size=10, name='Calibri')

                row += 1

        # Add visualization on the right side
        action_dist = metrics.get('action_distribution', {})
        if action_dist:
            try:
                chart_buffer = self.viz.create_action_distribution_chart(action_dist)
                img = XLImage(chart_buffer)
                img.width = 650
                img.height = 400
                # Place chart starting at column I (right side of the data table)
                ws.add_image(img, 'I5')
            except Exception as e:
                logger.warning(f"Could not create action distribution chart: {e}")

        # Add LLM Findings Section
        row_for_findings = row + 3 if rule_effectiveness else row + 1
        self._add_llm_findings_section(ws, row_for_findings, "LLM-Generated Rule Action Analysis Findings", merge_cols='A:G', findings=llm_findings)

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 2  # Gap
        ws.column_dimensions['I'].width = 2
        ws.column_dimensions['J'].width = 2
        ws.column_dimensions['K'].width = 2
        ws.column_dimensions['L'].width = 2
