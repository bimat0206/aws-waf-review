"""Traffic analysis sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class TrafficAnalysisSheet(BaseSheet):
    """Sheet generator for traffic analysis."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, metrics: Dict[str, Any]) -> None:
        """
        Create the Traffic Analysis sheet.

        This sheet analyzes traffic patterns and geographic distribution, including:
        - Daily traffic trends over time (requests, blocks, allows)
        - Geographic distribution of traffic by country
        - Threat scores by geographic location
        - Visual charts showing traffic patterns and trends

        Metrics explained:
        - Total Requests: Number of requests from each country
        - Blocked: Requests blocked by WAF rules
        - Allowed: Requests allowed through WAF
        - Threat Score: Percentage of requests blocked (higher = more threats)
        """
        logger.info("Creating Traffic Analysis sheet...")

        ws = self.workbook.create_sheet("Traffic Analysis")

        # Title with professional styling
        ws['A1'] = 'Traffic Analysis'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:E2')

        # Description
        ws['A3'] = 'Analyzes traffic patterns over time and geographic distribution to identify trends, peak periods, and regional threat sources.'
        ws['A3'].font = Font(size=10, italic=True, color='606060', name='Calibri')
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:E3')
        ws.row_dimensions[3].height = 30

        row = 5

        # Daily trends chart - will be placed on the right side later
        daily_chart_start_row = row
        daily_data = metrics.get('daily_trends')
        has_daily_chart = False

        if daily_data is not None and not daily_data.empty:
            has_daily_chart = True

        # Geographic distribution
        geo_data = metrics.get('geographic_distribution', [])
        if geo_data:
            ws[f'A{row}'] = 'Geographic Distribution'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            # Create table
            headers = ['Country', 'Total Requests', 'Blocked', 'Allowed', 'Threat Score']
            self._format_header_row(ws, row, headers)
            row += 1

            for idx, country_data in enumerate(geo_data[:20]):  # Top 20
                highlight = idx % 2 == 0
                threat_score = country_data.get('threat_score', 0)

                row_data = [
                    country_data.get('country', ''),
                    country_data.get('total_requests', 0),
                    country_data.get('blocked_requests', 0),
                    country_data.get('allowed_requests', 0),
                    f"{threat_score:.1f}%"
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                # Color code threat score
                threat_cell = ws.cell(row=row, column=5)
                if threat_score > 50:
                    threat_cell.fill = self.danger_fill
                    threat_cell.font = Font(bold=True, size=10, color='C00000', name='Calibri')
                elif threat_score > 25:
                    threat_cell.fill = self.warning_fill
                    threat_cell.font = Font(bold=True, size=10, color='FF8C00', name='Calibri')

                row += 1

        else:
            ws[f'A{row}'] = 'Geographic Distribution'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            ws[f'A{row}'] = 'No geographic data available.'
            ws[f'A{row}'].font = Font(italic=True, color='808080', name='Calibri')
            ws.merge_cells(f'A{row}:E{row}')
            row += 2

        # Add charts on the right side
        chart_col = 'G'  # Start charts at column G (right side)

        # Daily traffic chart
        if has_daily_chart:
            try:
                chart_buffer = self.viz.create_daily_traffic_chart(daily_data)
                img = XLImage(chart_buffer)
                img.width = 700
                img.height = 350
                ws.add_image(img, f'{chart_col}{daily_chart_start_row}')
            except Exception as e:
                logger.warning(f"Could not create daily traffic chart: {e}")

        # Geographic chart (if available)
        if geo_data:
            try:
                chart_buffer = self.viz.create_geographic_threat_chart(geo_data)
                img = XLImage(chart_buffer)
                img.width = 700
                img.height = 400
                # Position below daily chart (approximately 23 rows down)
                chart_row = daily_chart_start_row + 23 if has_daily_chart else daily_chart_start_row
                ws.add_image(img, f'{chart_col}{chart_row}')
            except Exception as e:
                logger.warning(f"Could not create geographic chart: {e}")

        # Add LLM Findings Section
        if geo_data or has_daily_chart:
            row_for_findings = row + 3 if geo_data else row + 1
        else:
            row_for_findings = row + 1

        self._add_llm_findings_section(ws, row_for_findings, "LLM-Generated Traffic Analysis Findings")

        # Auto-adjust columns
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions['F'].width = 2  # Gap
        ws.column_dimensions['G'].width = 2
        ws.column_dimensions['H'].width = 2
        ws.column_dimensions['I'].width = 2
        ws.column_dimensions['J'].width = 2
