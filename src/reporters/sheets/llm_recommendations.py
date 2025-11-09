"""LLM recommendations sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict, List, Optional
from openpyxl.styles import Alignment, Font, PatternFill
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class LLMRecommendationsSheet(BaseSheet):
    """Sheet generator for LLM-based security recommendations."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, llm_analysis: Optional[Dict[str, Any]] = None, analysis_metadata: Optional[Dict[str, Any]] = None) -> None:
        """
        Create the LLM Recommendations sheet.

        Args:
            llm_analysis: Parsed LLM analysis results. If None, creates manual template.
            analysis_metadata: Metadata about LLM analysis (model, tokens, cost, etc.)
        """
        if llm_analysis is None:
            logger.info("Creating LLM Recommendations sheet (manual template mode)...")
            return self._build_template()
        else:
            logger.info("Creating LLM Recommendations sheet (auto-populated mode)...")
            return self._build_with_recommendations(llm_analysis, analysis_metadata)

    def _build_template(self) -> None:
        """Create manual template for LLM recommendations."""
        ws = self.workbook.create_sheet("LLM Recommendations")

        # Title with professional styling
        ws['A1'] = 'AI-Generated Security Recommendations'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:D2')

        row = 4

        # Instructions box with highlighted background
        ws[f'A{row}'] = 'Instructions:'
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        instructions = [
            '1. A prompt file has been saved in output/ directory with all WAF metrics injected',
            '2. Copy the prompt and paste it into ChatGPT/Claude/Gemini',
            '3. Paste the AI-generated recommendations into the sections below',
            '4. Review and validate all recommendations before implementation'
        ]

        for instruction in instructions:
            ws[f'A{row}'] = instruction
            ws[f'A{row}'].font = Font(italic=True, size=10, name='Calibri')
            ws[f'A{row}'].fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

        # Sections with professional styling
        sections = [
            ('Critical Findings (Immediate Action Required)', 'Short-term implementation to address urgent vulnerabilities', 'FF6B6B'),
            ('Mid/Long-Term Initiatives', 'Sustained improvement and adaptability to evolving threats', 'FFA500'),
            ('Low Priority Suggestions (Nice to Have)', 'Nice to have improvements', '6BCF7F')
        ]

        for section_title, section_desc, color in sections:
            row += 2
            ws[f'A{row}'] = section_title
            ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'A{row}'].border = self.thin_border
            ws.merge_cells(f'A{row}:D{row}')
            ws.row_dimensions[row].height = 25
            row += 1

            ws[f'A{row}'] = section_desc
            ws[f'A{row}'].font = Font(italic=True, size=10, name='Calibri')
            ws[f'A{row}'].fill = self.highlight_fill
            ws.merge_cells(f'A{row}:D{row}')
            row += 1

            # Template rows
            headers = ['Priority', 'Recommendation', 'Impact', 'Action Items']
            self._format_header_row(ws, row, headers)
            row += 1

            # Add 3 empty rows with borders for each section
            for i in range(3):
                for col_idx in range(1, 5):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = self.thin_border
                    if i % 2 == 0:
                        cell.fill = self.highlight_fill
                row += 1

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50

    def _build_with_recommendations(self, analysis: Dict[str, Any], metadata: Optional[Dict[str, Any]]) -> None:
        """Create sheet with auto-populated LLM recommendations."""
        ws = self.workbook.create_sheet("LLM Recommendations")

        # Title with professional styling
        ws['A1'] = 'AI-Generated Security Recommendations'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:E2')

        row = 4

        # Analysis Metadata
        if metadata:
            ws[f'A{row}'] = 'Analysis Metadata'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            meta_items = [
                ('Provider', metadata.get('provider', 'N/A')),
                ('Model', metadata.get('model', 'N/A')),
                ('Total Tokens', f"{metadata.get('tokens_used', {}).get('total', 0):,}"),
                ('Estimated Cost', f"${metadata.get('cost_estimate', 0):.4f}"),
                ('Analysis Duration', f"{metadata.get('duration', 0):.2f}s")
            ]

            for idx, (label, value) in enumerate(meta_items):
                highlight = idx % 2 == 0
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'A{row}'].border = self.thin_border
                if highlight:
                    ws[f'A{row}'].fill = self.highlight_fill

                ws[f'B{row}'] = str(value)
                ws[f'B{row}'].font = self.data_font
                ws[f'B{row}'].border = self.thin_border
                if highlight:
                    ws[f'B{row}'].fill = self.highlight_fill
                row += 1

            row += 1

        # Executive Summary
        exec_summary = analysis.get('executive_summary', {})
        if exec_summary:
            ws[f'A{row}'] = 'Executive Summary'
            ws[f'A{row}'].font = self.subtitle_font
            ws.merge_cells(f'A{row}:E{row}')
            row += 1

            # Security Posture Assessment
            posture = exec_summary.get('security_posture', exec_summary.get('security_score', 'Medium'))

            # Convert numeric score to High/Medium/Low if needed
            if isinstance(posture, (int, float)):
                if posture >= 80:
                    posture = 'High'
                elif posture >= 60:
                    posture = 'Medium'
                else:
                    posture = 'Low'

            ws[f'A{row}'] = 'Security Posture Assessment'
            ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
            ws[f'A{row}'].border = self.thin_border

            ws[f'B{row}'] = str(posture).upper()
            ws[f'B{row}'].font = Font(bold=True, size=14, name='Calibri')
            ws[f'B{row}'].border = self.thin_border

            # Color code the assessment
            if str(posture).upper() == 'HIGH':
                ws[f'B{row}'].font = Font(bold=True, size=14, color='008000', name='Calibri')
                ws[f'B{row}'].fill = self.success_fill
            elif str(posture).upper() == 'MEDIUM':
                ws[f'B{row}'].font = Font(bold=True, size=14, color='FF8C00', name='Calibri')
                ws[f'B{row}'].fill = self.warning_fill
            else:
                ws[f'B{row}'].font = Font(bold=True, size=14, color='C00000', name='Calibri')
                ws[f'B{row}'].fill = self.danger_fill

            row += 1

            # Assessment Breakdown
            breakdown = exec_summary.get('assessment_breakdown', {})
            if breakdown:
                ws[f'A{row}'] = 'Assessment Breakdown'
                ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'A{row}'].border = self.thin_border
                ws[f'A{row}'].fill = self.highlight_fill
                ws[f'B{row}'].border = self.thin_border
                ws[f'B{row}'].fill = self.highlight_fill
                ws.merge_cells(f'B{row}:E{row}')
                row += 1

                # Breakdown items
                breakdown_items = [
                    ('Rule Coverage', breakdown.get('rule_coverage', 'Medium')),
                    ('Threat Detection', breakdown.get('threat_detection', 'Medium')),
                    ('Logging & Monitoring', breakdown.get('logging_monitoring', 'Medium')),
                    ('Configuration Security', breakdown.get('configuration_security', 'Medium')),
                    ('Response Readiness', breakdown.get('response_readiness', 'Medium'))
                ]

                for idx, (label, value) in enumerate(breakdown_items):
                    ws[f'A{row}'] = f'  • {label}'
                    ws[f'A{row}'].font = Font(size=10, name='Calibri')
                    ws[f'A{row}'].border = self.thin_border

                    ws[f'B{row}'] = str(value).upper()
                    ws[f'B{row}'].font = Font(bold=True, size=10, name='Calibri')
                    ws[f'B{row}'].border = self.thin_border

                    # Color code each breakdown item
                    if str(value).upper() == 'HIGH':
                        ws[f'B{row}'].font = Font(bold=True, size=10, color='008000', name='Calibri')
                    elif str(value).upper() == 'MEDIUM':
                        ws[f'B{row}'].font = Font(bold=True, size=10, color='FF8C00', name='Calibri')
                    else:
                        ws[f'B{row}'].font = Font(bold=True, size=10, color='C00000', name='Calibri')

                    row += 1

            # Overall Assessment text
            if exec_summary.get('assessment'):
                ws[f'A{row}'] = 'Overall Assessment'
                ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
                ws[f'B{row}'] = exec_summary['assessment']
                ws[f'B{row}'].font = self.data_font
                ws[f'B{row}'].alignment = Alignment(wrap_text=True)
                ws.merge_cells(f'B{row}:E{row}')
                ws.row_dimensions[row].height = 40
                row += 1

            row += 1

        # Findings Sections with New Professional Format
        findings_sections = [
            (
                'Critical Findings (Immediate Action Required)',
                analysis.get('critical_findings', []),
                'FF6B6B',
                'Critical',
                None
            ),
            (
                'Mid/Long-Term Recommendations',
                analysis.get('high_priority', []) + analysis.get('medium_priority', []),
                'FFA500',
                'Strategic',
                'These initiatives require a more strategic approach and ongoing effort to maintain and improve the WAF\'s effectiveness over time'
            ),
            (
                'Low Priority Suggestions (Nice to Have)',
                analysis.get('low_priority', []),
                '6BCF7F',
                'Low',
                'Nice to have improvements'
            )
        ]

        for section_title, findings, color, priority, subtitle in findings_sections:
            if not findings:
                continue

            # Section header
            ws[f'A{row}'] = section_title
            ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'A{row}'].border = self.thin_border
            ws.merge_cells(f'A{row}:E{row}')
            ws.row_dimensions[row].height = 25
            row += 1

            # Optional subtitle
            if subtitle:
                ws[f'A{row}'] = subtitle
                ws[f'A{row}'].font = Font(italic=True, size=10, color='666666', name='Calibri')
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'A{row}:E{row}')
                row += 1

            # Table headers - NEW FORMAT
            headers = ['No', 'Recommendation', 'Expected Impact', 'Action Items', 'Rationale']
            self._format_header_row(ws, row, headers)
            row += 1

            # Recommendations data
            for idx, finding in enumerate(findings, start=1):
                highlight = idx % 2 == 0

                # Format action items as bullet list if it's a list
                action_items = finding.get('action', finding.get('actions', ''))
                if isinstance(action_items, list):
                    action_items_formatted = '\n'.join([f'• {item}' for item in action_items])
                else:
                    # Try to convert string with newlines to bullet list
                    if '\n' in str(action_items):
                        items = [line.strip() for line in str(action_items).split('\n') if line.strip()]
                        action_items_formatted = '\n'.join([f'• {item}' for item in items])
                    else:
                        action_items_formatted = f'• {action_items}' if action_items else ''

                row_data = [
                    str(idx),  # No column
                    finding.get('title', finding.get('finding', finding.get('recommendation', ''))),  # Recommendation
                    finding.get('impact', finding.get('expected_impact', '')),  # Expected Impact
                    action_items_formatted,  # Action Items (bullet list)
                    finding.get('rationale', finding.get('reason', ''))  # Rationale
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx)
                    self._format_data_cell(cell, value, highlight)

                    # Set specific column formatting
                    if col_idx == 1:  # No column - center align
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif col_idx in [2, 3, 4, 5]:  # Text columns - wrap text
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

                # Adjust row height based on content
                ws.row_dimensions[row].height = max(40, min(len(str(row_data[3])) // 40, 120))
                row += 1

            row += 2

        # Auto-adjust columns for new format
        ws.column_dimensions['A'].width = 6   # No
        ws.column_dimensions['B'].width = 35  # Recommendation
        ws.column_dimensions['C'].width = 25  # Expected Impact
        ws.column_dimensions['D'].width = 40  # Action Items
        ws.column_dimensions['E'].width = 30  # Rationale
