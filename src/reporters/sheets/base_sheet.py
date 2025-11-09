"""
Base Sheet Class

Provides common styling, formatting, and helper methods for all Excel sheet generators.
"""

import logging
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


class BaseSheet:
    """
    Base class for all sheet generators.
    Provides common styling and formatting methods.
    """

    def __init__(self):
        """Initialize common styling properties."""
        # Professional Styling Theme
        self.header_font = Font(bold=True, size=11, color='FFFFFF', name='Calibri')
        self.header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # Professional blue
        self.title_font = Font(bold=True, size=14, color='1F4E78', name='Calibri')
        self.subtitle_font = Font(bold=True, size=12, color='2C3E50', name='Calibri')
        self.data_font = Font(size=10, name='Calibri')
        self.highlight_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')  # Light gray
        self.success_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light green
        self.warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Light yellow
        self.danger_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Light red

        # Border styles
        self.thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        self.thick_border = Border(
            left=Side(style='medium', color='1F4E78'),
            right=Side(style='medium', color='1F4E78'),
            top=Side(style='medium', color='1F4E78'),
            bottom=Side(style='medium', color='1F4E78')
        )
        self.viz = None

    def _apply_cell_style(self, cell, font=None, fill=None, border=None, alignment=None):
        """
        Apply consistent styling to a cell.

        Args:
            cell: The cell to style
            font: Font object to apply
            fill: PatternFill object to apply
            border: Border object to apply
            alignment: Alignment object to apply
        """
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment

    def _format_header_row(self, ws, row, columns, start_col=1):
        """
        Format a header row with professional styling.

        Args:
            ws: Worksheet object
            row: Row number
            columns: List of column headers
            start_col: Starting column number (default: 1)
        """
        for col_idx, header in enumerate(columns, start=start_col):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            self._apply_cell_style(
                cell,
                font=self.header_font,
                fill=self.header_fill,
                border=self.thin_border,
                alignment=Alignment(horizontal='center', vertical='center', wrap_text=True)
            )
        ws.row_dimensions[row].height = 25

    def _format_data_cell(self, cell, value, highlight=False):
        """
        Format a data cell with professional styling.

        Args:
            cell: The cell to format
            value: The value to set
            highlight: Whether to apply highlight fill (default: False)
        """
        cell.value = value
        cell.font = self.data_font
        cell.border = self.thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=False)
        # Only apply fill if highlight is True
        if highlight:
            cell.fill = self.highlight_fill

    def _format_data_rows_bulk(self, ws, start_row, data_rows, start_col=1, alternating_highlight=True):
        """
        Format multiple data rows efficiently in bulk.

        Args:
            ws: Worksheet object
            start_row: Starting row number
            data_rows: List of lists containing row data
            start_col: Starting column number (default: 1)
            alternating_highlight: Whether to apply alternating row highlighting (default: True)
        """
        for row_idx, row_data in enumerate(data_rows):
            row_num = start_row + row_idx
            highlight = alternating_highlight and row_idx % 2 == 0
            
            # Use append() for bulk insertion if at the end of the sheet
            if row_num == ws.max_row + 1:
                ws.append(row_data)
                actual_row = ws.max_row  # Get the actual row number after append
            else:
                # Insert at specific row if not at the end
                for col_idx, value in enumerate(row_data, start=start_col):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = value
                    cell.font = self.data_font
                    cell.border = self.thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=False)
                    if highlight:
                        cell.fill = self.highlight_fill
            # Apply highlighting to the entire row after insertion
            if highlight and row_num != ws.max_row:  # Only if not using append
                for col_idx in range(start_col, start_col + len(row_data)):
                    cell = ws.cell(row=row_num, column=col_idx)
                    if not cell.fill or cell.fill == self.highlight_fill:
                        cell.fill = self.highlight_fill

    def _add_sheet_title(self, ws, title, row=1, merge_range='A1:D1'):
        """
        Add a professional title to a sheet.

        Args:
            ws: Worksheet object
            title: Title text
            row: Row number (default: 1)
            merge_range: Cell range to merge (default: 'A1:D1')
        """
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(merge_range)
        ws.row_dimensions[row].height = 30

    def _add_sheet_subtitle(self, ws, row=2, merge_range='A2:D2'):
        """
        Add a generated timestamp subtitle to a sheet.

        Args:
            ws: Worksheet object
            row: Row number (default: 2)
            merge_range: Cell range to merge (default: 'A2:D2')
        """
        ws[f'A{row}'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws[f'A{row}'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells(merge_range)

    def _add_section_header(self, ws, row, title, merge_range):
        """
        Add a section header with subtitle styling.

        Args:
            ws: Worksheet object
            row: Row number
            title: Section title
            merge_range: Cell range to merge
        """
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = self.subtitle_font
        ws.merge_cells(merge_range)

    def _add_llm_findings_section(self, ws, start_row, section_title="LLM-Generated Findings", merge_cols='A:E', findings=None):
        """
        Add an LLM-generated findings section with template or populated data.

        Args:
            ws: Worksheet object
            start_row: Starting row number
            section_title: Title for the findings section
            merge_cols: Column range for merging (default: 'A:E')
            findings: Optional list of finding dicts with keys: finding, severity, recommendation

        Returns:
            int: Next available row number after the section
        """
        row = start_row

        # Section header
        ws[f'A{row}'] = section_title
        ws[f'A{row}'].font = Font(bold=True, size=14, color='1F4E78', name='Calibri')
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        merge_range = f'A{row}:{merge_cols.split(":")[1]}{row}'
        ws.merge_cells(merge_range)
        ws.row_dimensions[row].height = 25
        row += 1

        # Instructions (only show in template mode)
        if not findings or len(findings) == 0:
            ws[f'A{row}'] = 'Instructions: This section will be automatically populated when LLM analysis is run (Option 6 in main menu).'
            ws[f'A{row}'].font = Font(italic=True, size=10, color='666666', name='Calibri')
            merge_range = f'A{row}:{merge_cols.split(":")[1]}{row}'
            ws.merge_cells(merge_range)
            row += 1

        # Template table
        headers = ['No', 'Finding', 'Severity', 'Recommendation']
        self._format_header_row(ws, row, headers)
        row += 1

        # Populate with findings or add template rows
        if findings and len(findings) > 0:
            # Populated mode - add actual findings
            for idx, finding in enumerate(findings, start=1):
                # Column 1: Number
                cell = ws.cell(row=row, column=1)
                cell.value = idx
                cell.font = Font(bold=True, size=10, name='Calibri')
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Column 2: Finding (wrap text)
                cell = ws.cell(row=row, column=2)
                cell.value = finding.get('finding', '')
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Column 3: Severity (color-coded)
                severity = finding.get('severity', 'MEDIUM').upper()
                cell = ws.cell(row=row, column=3)
                cell.value = severity
                cell.font = Font(bold=True, size=10, name='Calibri')
                cell.border = self.thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Color-code severity
                if severity == 'HIGH':
                    cell.fill = self.danger_fill  # Red
                elif severity == 'MEDIUM':
                    cell.fill = self.warning_fill  # Yellow
                elif severity == 'LOW':
                    cell.fill = self.success_fill  # Green

                # Column 4: Recommendation (wrap text)
                cell = ws.cell(row=row, column=4)
                cell.value = finding.get('recommendation', '')
                cell.font = self.data_font
                cell.border = self.thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Set row height for wrapped text
                ws.row_dimensions[row].height = 60
                row += 1
        else:
            # Template mode - add 3 empty rows
            for i in range(3):
                highlight = i % 2 == 0
                for col_idx in range(1, 5):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = self.thin_border
                    if highlight:
                        cell.fill = self.highlight_fill
                row += 1

        return row

    def create(self, workbook, *args, **kwargs):
        """
        Create the sheet. Must be implemented by subclasses.

        Args:
            workbook: The workbook to add the sheet to
            *args: Additional positional arguments
            **kwargs: Additional keyword arguments

        Raises:
            NotImplementedError: This method must be implemented by subclasses
        """
        raise NotImplementedError("Subclasses must implement the create() method")
