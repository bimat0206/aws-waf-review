"""Rule effectiveness sheet generator."""

import logging
from datetime import datetime
from typing import Any, Dict
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from .base_sheet import BaseSheet

logger = logging.getLogger(__name__)


class RuleEffectivenessSheet(BaseSheet):
    """Sheet generator for rule effectiveness."""

    def __init__(self, workbook):
        super().__init__()
        self.workbook = workbook

    def build(self, metrics: Dict[str, Any]) -> None:
        """
        Create the Rule Effectiveness sheet.

        This sheet evaluates the performance and effectiveness of WAF rules, including:
        - Rule hit counts and frequency
        - Block vs. Allow ratios per rule
        - Rule types (REGULAR, MANAGED, RATE_BASED)
        - Attack type distribution based on rules triggered
        - Visual charts showing rule performance

        Metrics explained:
        - Hit Count: Total number of times the rule was triggered
        - Blocks: Number of requests blocked by this rule
        - Allows: Number of requests allowed by this rule
        - Hit Rate %: Percentage of total requests that triggered this rule
        - Block Rate %: Percentage of rule hits that resulted in blocks
          - Green (>0%): Rule is actively blocking threats
          - Red (0%): Rule never blocks (may need review)
        """
        logger.info("Creating Rule Effectiveness sheet...")

        ws = self.workbook.create_sheet("Rule Effectiveness")

        # Title with professional styling
        ws['A1'] = 'Rule Effectiveness Analysis'
        ws['A1'].font = Font(bold=True, size=18, color='1F4E78', name='Calibri')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 30

        # Subtitle
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True, color='808080', name='Calibri')
        ws.merge_cells('A2:G2')

        # Description
        ws['A3'] = 'Analyzes WAF rule performance to identify effective rules, unused rules, and optimization opportunities.'
        ws['A3'].font = Font(size=10, italic=True, color='606060', name='Calibri')
        ws['A3'].alignment = Alignment(wrap_text=True)
        ws.merge_cells('A3:G3')
        ws.row_dimensions[3].height = 30

        row = 5

        # Rule effectiveness table
        rule_data = metrics.get('rule_effectiveness', [])
        if rule_data:
            # Headers
            headers = ['Rule ID', 'Rule Type', 'Hit Count', 'Blocks', 'Allows', 'Hit Rate %', 'Block Rate %']
            self._format_header_row(ws, row, headers)
            row += 1

            # Prepare data for bulk insertion
            data_rows = []
            for idx, rule in enumerate(rule_data):
                hit_rate = rule.get('hit_rate_percent', 0)

                row_data = [
                    rule.get('rule_id', '')[:50],
                    rule.get('rule_type', ''),
                    rule.get('hit_count', 0),
                    rule.get('blocks', 0),
                    rule.get('allows', 0),
                    f"{hit_rate:.1f}",
                    f"{rule.get('block_rate_percent', 0):.1f}"
                ]
                data_rows.append(row_data)

            # Insert data in bulk with alternating highlighting
            start_data_row = row
            self._format_data_rows_bulk(ws, start_data_row, data_rows, start_col=1, alternating_highlight=True)
            
            # Apply special formatting for hit rate column (column 6)
            for idx in range(len(rule_data)):
                hit_rate = rule_data[idx].get('hit_rate_percent', 0)
                hit_rate_cell = ws.cell(row=start_data_row + idx, column=6)
                if hit_rate == 0:
                    hit_rate_cell.fill = self.danger_fill
                    hit_rate_cell.font = Font(bold=True, size=10, color='C00000', name='Calibri')
                elif hit_rate > 10:
                    hit_rate_cell.fill = self.success_fill
                    hit_rate_cell.font = Font(bold=True, size=10, color='008000', name='Calibri')
                    
            row += len(rule_data)

        # Add charts on the right side
        chart_col = 'I'  # Start charts at column I (right side, after columns A-G)
        chart_start_row = 5

        # Rule effectiveness chart
        rule_data = metrics.get('rule_effectiveness', [])
        if rule_data:
            try:
                chart_buffer = self.viz.create_rule_effectiveness_chart(rule_data)
                img = XLImage(chart_buffer)
                img.width = 650
                img.height = 400
                ws.add_image(img, f'{chart_col}{chart_start_row}')
            except Exception as e:
                logger.warning(f"Could not create rule effectiveness chart: {e}")

        # Attack type distribution chart
        attack_data = metrics.get('attack_type_distribution', {})
        if attack_data:
            try:
                chart_buffer = self.viz.create_attack_type_chart(attack_data)
                img = XLImage(chart_buffer)
                img.width = 650
                img.height = 400
                # Position below rule effectiveness chart (approximately 26 rows down)
                attack_chart_row = chart_start_row + 26
                ws.add_image(img, f'{chart_col}{attack_chart_row}')
            except Exception as e:
                logger.warning(f"Could not create attack type chart: {e}")

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 50
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['H'].width = 2  # Gap
        ws.column_dimensions['I'].width = 2
        ws.column_dimensions['J'].width = 2
        ws.column_dimensions['K'].width = 2
        ws.column_dimensions['L'].width = 2
